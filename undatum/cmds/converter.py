"""File format conversion module."""

import csv
import json
import logging
import xml.etree.ElementTree as etree
from collections import defaultdict

import bson
import orjson
import pandas
from bson import ObjectId
from xlrd import open_workbook as load_xls

from ..common.command_utils import ITERABLE_OPTIONS_KEYS, get_iterable_options  # noqa: F401
from ..common.errors import (
    FileNotFoundError,
    FormatError,
    PermissionError,
    ValidationError,
    find_similar_files,
)
from ..common.path_utils import validate_file_path
from ..constants import COMPRESSED_FILE_TYPES, SUPPORTED_FILE_TYPES
from ..utils import get_file_type, get_option

# Preferred order for suggesting writable conversion targets. The list is
# filtered against iterabledata's actual write capabilities at runtime so the
# suggestions can never contradict the read-only check below.
_PREFERRED_WRITABLE_FORMATS = ["csv", "jsonl", "json", "parquet", "orc", "avro", "bson"]


def _writable_format_suggestions() -> list[str]:
    """Return the preferred writable formats that iterabledata can actually write.

    Falls back to the full preferred list if capability reporting is unavailable.
    """
    try:
        from iterable.helpers.capabilities import supports_write
    except Exception:  # noqa: BLE001 - capabilities unavailable: use static list
        return list(_PREFERRED_WRITABLE_FORMATS)
    writable = []
    for fmt in _PREFERRED_WRITABLE_FORMATS:
        try:
            if supports_write(fmt) is not False:
                writable.append(fmt)
        except Exception:  # noqa: BLE001 - unknown format: skip it
            continue
    return writable or list(_PREFERRED_WRITABLE_FORMATS)


# Formats commonly suggested as writable conversion targets (capability-filtered).
WRITABLE_FORMAT_SUGGESTIONS = _writable_format_suggestions()

# Formats that report as writable but require an externally-supplied schema or
# message class to serialize (e.g. a compiled protobuf message). They cannot be
# used as generic conversion targets, so we fail fast with an actionable error
# instead of surfacing the engine's "requires '<x>' parameter" error. Keys are
# resolved output format ids/extensions; values name the required input.
_SCHEMA_REQUIRED_FORMATS: dict[str, str] = {
    "pb": "a compiled protobuf message class (message_class)",
    "protobuf": "a compiled protobuf message class (message_class)",
    "capnp": "a Cap'n Proto schema (schema_file and schema_name)",
    "thrift": "a generated Thrift struct class (struct_class)",
}

DEFAULT_BATCH_SIZE = 50000


PREFIX_STRIP = True
PREFIX = ""

LINEEND = b"\n"


def df_to_pyorc_schema(df):
    """Extracts column information from pandas dataframe and generate pyorc schema"""
    struct_schema = []
    for k, v in df.dtypes.to_dict().items():
        v = str(v)
        if v == "float64":
            struct_schema.append(f"{k}:float")
        elif v == "float32":
            struct_schema.append(f"{k}:float")
        elif v == "datetime64[ns]":
            struct_schema.append(f"{k}:timestamp")
        elif v == "int32":
            struct_schema.append(f"{k}:int")
        elif v == "int64":
            struct_schema.append(f"{k}:int")
        else:
            struct_schema.append(f"{k}:string")
    return struct_schema


def __copy_options(user_options, default_options):
    """If user provided option so we use it, if not, default option value should be used"""
    for k in default_options.keys():
        if k not in user_options.keys():
            user_options[k] = default_options[k]
    return user_options


def etree_to_dict(t, prefix_strip=True):
    """Convert XML element tree to dictionary."""
    tag = t.tag if not prefix_strip else t.tag.rsplit("}", 1)[-1]
    d = {tag: {} if t.attrib else None}
    children = list(t)
    if children:
        dd = defaultdict(list)
        for dc in map(etree_to_dict, children):
            for k, v in dc.items():
                if prefix_strip:
                    # Remove XML namespace prefix (e.g., '{http://...}tagname' -> 'tagname')
                    k = k.rsplit("}", 1)[-1]
                dd[k].append(v)
        d = {tag: {k: v[0] if len(v) == 1 else v for k, v in dd.items()}}
    if t.attrib:
        d[tag].update(("@" + k.rsplit("}", 1)[-1], v) for k, v in t.attrib.items())
    if t.text:
        text = t.text.strip()
        if children or t.attrib:
            tag = tag.rsplit("}", 1)[-1]
            if text:
                d[tag]["#text"] = text
        else:
            d[tag] = text
    return d


def xml_to_jsonl(fromname, toname, options=None, default_options=None):
    """Convert XML file to JSONL format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {"prefix_strip": True}
    options = __copy_options(options, default_options)
    with open(fromname, "rb") as ins, open(toname, "wb") as outf:
        n = 0
        for _event, elem in etree.iterparse(ins):
            shorttag = elem.tag.rsplit("}", 1)[-1]
            if shorttag == options["tagname"]:
                n += 1
                if options["prefix_strip"]:
                    j = etree_to_dict(elem, prefix_strip=options["prefix_strip"])
                else:
                    j = etree_to_dict(elem)
                outf.write(orjson.dumps(j[shorttag]))
                outf.write(LINEEND)
            if n % 500 == 0:
                logging.info("xml2jsonl: processed %d xml tags", n)
        logging.info("xml2jsonl: processed %d xml tags finally", n)


def xls_to_csv(fromname, toname, options=None, default_options=None):
    """Convert XLS file to CSV format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {
            "start_line": 0,
            "skip_end_rows": 0,
            "delimiter": ",",
            "encoding": "utf8",
        }
    options = __copy_options(options, default_options)
    b = load_xls(fromname)
    s = b.sheet_by_index(0)
    with open(toname, "w", encoding=options["encoding"]) as bc:
        bcw = csv.writer(bc, delimiter=options["delimiter"])
        n = 0
        end_row = s.nrows - options["skip_end_rows"]
        for row in range(options["start_line"], end_row):
            n += 1
            this_row = []
            for col in range(s.ncols):
                v = str(s.cell_value(row, col))
                v = v.replace("\n", " ").strip()
                this_row.append(v)
            bcw.writerow(this_row)
            if n % 10000 == 0:
                logging.info("xls2csv: processed %d records", n)


def csv_to_bson(fromname, toname, options=None, default_options=None):
    """Convert CSV file to BSON format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {"encoding": "utf8", "delimiter": ","}
    options = __copy_options(options, default_options)
    with open(fromname, encoding=options["encoding"]) as source:
        reader = csv.DictReader(source, delimiter=options["delimiter"])
        with open(toname, "wb") as output:
            n = 0
            for j in reader:
                n += 1
                rec = bson.BSON.encode(j)
                output.write(rec)
                if n % 10000 == 0:
                    logging.info("csv2bson: processed %d records", n)


def csv_to_jsonl(fromname, toname, options=None, default_options=None):
    """Convert CSV file to JSONL format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {"encoding": "utf8", "delimiter": ","}
    options = __copy_options(options, default_options)
    with open(fromname, encoding=options["encoding"]) as source:
        reader = csv.DictReader(source, delimiter=options["delimiter"])
        with open(toname, "wb") as output:
            n = 0
            for j in reader:
                n += 1
                output.write(json.dumps(j, ensure_ascii=False).encode("utf8"))
                output.write(b"\n")
                if n % 10000 == 0:
                    logging.info("csv2jsonl: processed %d records", n)


def xls_to_jsonl(fromname, toname, options=None, default_options=None):
    """Convert XLS file to JSONL format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {"start_page": 0, "start_line": 0, "fields": None}
    options = __copy_options(options, default_options)
    source = load_xls(fromname)
    sheet = source.sheet_by_index(options["start_page"])
    with open(toname, "wb") as output:
        n = 0
        fields = options["fields"].split(",") if options["fields"] is not None else None
        for rownum in range(options["start_line"], sheet.nrows):
            n += 1
            tmp = []
            for i in range(0, sheet.ncols):
                tmp.append(sheet.row_values(rownum)[i])
            if n == 1 and fields is None:
                fields = tmp
                continue
            line = orjson.dumps(dict(zip(fields, tmp)))
            output.write(line + LINEEND)
            if n % 10000 == 0:
                logging.info("xls2jsonl: processed %d records", n)


def xlsx_to_jsonl(fromname, toname, options=None, default_options=None):
    """Convert XLSX file to JSONL format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {"start_page": 0, "start_line": 0}
    from openpyxl import load_workbook as load_xlsx

    options = __copy_options(options, default_options)
    source = load_xlsx(fromname)
    # Use start_page to select the correct worksheet
    start_page = options.get("start_page", 0)
    if start_page >= len(source.worksheets):
        raise ValueError(
            f"start_page {start_page} exceeds available worksheets ({len(source.worksheets)})"
        )
    sheet = source.worksheets[start_page]
    with open(toname, "wb") as output:
        n = 0
        fields = options["fields"].split(",") if options["fields"] is not None else None
        for row in sheet.iter_rows():
            n += 1
            if n < options["start_line"]:
                continue
            tmp = []

            for cell in row:
                tmp.append(cell.value)
            if n == 1 and fields is None:
                fields = tmp
                continue
            line = orjson.dumps(dict(zip(fields, tmp)))
            output.write(line)
            output.write(LINEEND)
            if n % 10000 == 0:
                logging.debug("xlsx2bson: processed %d records", n)
    source.close()


def xlsx_to_bson(fromname, toname, options=None, default_options=None):
    """Convert XLSX file to BSON format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {"start_page": 0, "start_line": 0}
    from openpyxl import load_workbook as load_xlsx

    options = __copy_options(options, default_options)
    source = load_xlsx(fromname)
    sheet = source.active  # FIXME! Use start_page instead
    with open(toname, "wb") as output:
        n = 0
        fields = options["fields"].split(",") if options["fields"] is not None else None
        for row in sheet.iter_rows():
            n += 1
            if n < options["start_line"]:
                continue
            tmp = []

            for cell in row:
                tmp.append(cell.value)
            if n == 1 and fields is None:
                fields = tmp
                continue
            output.write(bson.BSON.encode(dict(zip(fields, tmp))))

            if n % 10000 == 0:
                logging.debug("xlsx2bson: processed %d records", n)
    source.close()


def xls_to_bson(fromname, toname, options=None, default_options=None):
    """Convert XLS file to BSON format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {"start_page": 0, "start_line": 0}
    options = __copy_options(options, default_options)
    source = load_xls(fromname)
    sheet = source.sheet_by_index(options["start_page"])
    with open(toname, "wb") as output:
        n = 0
        for rownum in range(options["start_line"], sheet.nrows):
            n += 1
            tmp = []
            for i in range(0, sheet.ncols):
                tmp.append(sheet.row_values(rownum)[i])
            output.write(bson.BSON.encode(dict(zip(options["fields"], tmp))))
            if n % 10000 == 0:
                logging.info("xls2bson: processed %d records", n)


def _is_flat(item):
    """Check if dictionary item is flat (no nested structures)."""
    for _k, v in item.items():
        if isinstance(v, (dict, tuple, list)):
            return False
    return True


def express_analyze_jsonl(filename, itemlimit=100):
    """Quickly analyze JSONL file structure."""
    isflat = True
    n = 0
    keys = set()
    with open(filename, encoding="utf8") as f:
        for line in f:
            n += 1
            if n > itemlimit:
                break
            record = orjson.loads(line)
            if isflat:
                if not _is_flat(record):
                    isflat = False
            if len(keys) == 0:
                keys = set(record.keys())
            else:
                keys = keys.union(set(record.keys()))
    keys = list(keys)
    keys.sort()
    return {"isflat": isflat, "keys": keys}


def jsonl_to_csv(fromname, toname, options=None, default_options=None):
    """Convert JSONL file to CSV format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {"force_flat": False, "useitems": 100, "delimiter": ","}
    options = __copy_options(options, default_options)
    analysis = express_analyze_jsonl(fromname, itemlimit=options["useitems"])
    if not options["force_flat"] and not analysis["isflat"]:
        logging.error(
            "File %s is not flat and 'force_flat' flag not set. File not converted", fromname
        )
        return
    keys = analysis["keys"]
    with open(toname, "w", encoding="utf8") as out:
        writer = csv.writer(out, delimiter=options["delimiter"])
        writer.writerow(keys)
        with open(fromname, encoding="utf8") as f:
            n = 0
            for line in f:
                n += 1
                record = orjson.loads(line)
                item = []
                for k in keys:
                    if k in record:
                        item.append(record[k])
                    else:
                        item.append("")
                writer.writerow(item)
                if n % 10000 == 0:
                    logging.info("jsonl2csv: processed %d records", n)


def default(obj):
    """Default serializer for BSON ObjectId."""
    if isinstance(obj, ObjectId):
        return str(obj)
    return None


def bson_to_jsonl(fromname, toname, options=None, default_options=None):
    """Convert BSON file to JSONL format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {}
    options = __copy_options(options, default_options)
    with open(fromname, "rb") as source:
        with open(toname, "wb") as output:
            n = 0
            for r in bson.decode_file_iter(source):
                n += 1
                output.write(orjson.dumps(r, default=default))
                output.write(LINEEND)
                if n % 10000 == 0:
                    logging.info("bson2jsonl: processed %d records", n)


def json_to_jsonl(fromname, toname, options=None, default_options=None):
    """Simple implementation of JSON to JSON lines conversion.

    Assumes that JSON is an array or dict with 1st level value with data.
    """
    if options is None:
        options = {}
    if default_options is None:
        default_options = {}
    options = __copy_options(options, default_options)
    source = open(fromname, "rb")
    source_data = json.load(source)
    data = source_data
    if "tagname" in options.keys():
        if isinstance(source_data, dict) and options["tagname"] in source_data:
            data = data[options["tagname"]]
    with open(toname, "wb") as output:
        n = 0
        for r in data:
            n += 1
            output.write(orjson.dumps(r) + LINEEND)
            if n % 10000 == 0:
                logging.info("json2jsonl: processed %d records", n)
    source.close()


def csv_to_parquet(fromname, toname, options=None, default_options=None):
    """Convert CSV file to Parquet format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {"encoding": "utf8", "delimiter": ",", "compression": "brotli"}
    options = __copy_options(options, default_options)
    df = pandas.read_csv(fromname, delimiter=options["delimiter"], encoding=options["encoding"])
    comp = options["compression"] if options["compression"] != "None" else None
    df.to_parquet(toname, compression=comp)


def jsonl_to_parquet(fromname, toname, options=None, default_options=None):
    """Convert JSONL file to Parquet format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {"force_flat": False, "useitems": 100, "compression": "brotli"}
    options = __copy_options(options, default_options)
    df = pandas.read_json(fromname, lines=True, encoding=options["encoding"])
    comp = options["compression"] if options["compression"] != "None" else None
    df.to_parquet(toname, compression=comp)


PYORC_COMPRESSION_MAP = {"zstd": 5, "snappy": 2, "zlib": 1, "lzo": 3, "lz4": 4, "None": 0}


def csv_to_orc(fromname, toname, options=None, default_options=None):
    """Converts CSV file to ORC file."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {"encoding": "utf8", "delimiter": ",", "compression": "zstd"}
    import pyorc

    options = __copy_options(options, default_options)
    comp_key = options["compression"]
    compression = PYORC_COMPRESSION_MAP[comp_key] if comp_key in PYORC_COMPRESSION_MAP.keys() else 0
    with open(fromname, encoding=options["encoding"]) as source:
        reader = csv.DictReader(source, delimiter=options["delimiter"])
        struct_schema = []
        for field in reader.fieldnames:
            struct_schema.append(f"{field}:string")
        schema_str = ",".join(struct_schema)
        with open(toname, "wb") as output:
            writer = pyorc.Writer(
                output,
                f"struct<{schema_str}>",
                struct_repr=pyorc.StructRepr.DICT,
                compression=compression,
                compression_strategy=1,
            )
            n = 0
            for row in reader:
                n += 1
                try:
                    writer.write(row)
                except TypeError:
                    print("Error processing row %d. Skip and continue", n)


def jsonl_to_orc(fromname, toname, options=None, default_options=None):
    """Converts JSON file to ORC file."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {"force_flat": False, "useitems": 100, "compression": "zstd"}
    import pyorc

    options = __copy_options(options, default_options)
    comp_key = options["compression"]
    compression = PYORC_COMPRESSION_MAP[comp_key] if comp_key in PYORC_COMPRESSION_MAP.keys() else 0
    df = pandas.read_json(fromname, lines=True, encoding=options["encoding"])
    df.info()
    struct_schema = df_to_pyorc_schema(df)
    schema_str = ",".join(struct_schema)
    with open(toname, "wb") as output:
        writer = pyorc.Writer(
            output,
            f"struct<{schema_str}>",
            struct_repr=pyorc.StructRepr.DICT,
            compression=compression,
            compression_strategy=1,
        )
        writer.writerows(df.to_dict(orient="records"))


def csv_to_avro(fromname, toname, options=None, default_options=None):
    """Converts CSV file to AVRO file."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {"encoding": "utf8", "delimiter": ",", "compression": "deflate"}
    import avro.schema
    from avro.datafile import DataFileWriter
    from avro.io import DatumWriter

    options = __copy_options(options, default_options)
    with open(fromname, encoding=options["encoding"]) as source:
        reader = csv.DictReader(source, delimiter=options["delimiter"])

        schema_dict = {"namespace": "data.avro", "type": "record", "name": "Record", "fields": []}

        for field in reader.fieldnames:
            schema_dict["fields"].append({"name": field, "type": "string"})
        schema = avro.schema.parse(json.dumps(schema_dict))
        with open(toname, "wb") as output:
            writer = DataFileWriter(output, DatumWriter(), schema, codec=options["compression"])
            n = 0
            for row in reader:
                n += 1
                try:
                    writer.append(row)
                except TypeError:
                    print("Error processing row %d. Skip and continue", n)


DEFAULT_HEADERS_DETECT_LIMIT = 1000


def make_flat(item):
    """Flatten nested structures in dictionary by converting to strings."""
    result = {}
    for k, v in item.items():
        if isinstance(v, (tuple, list, dict)):
            result[k] = str(v)
        else:
            result[k] = v
    return result


class Converter:
    """File format converter handler."""

    def __init__(self, batch_size=DEFAULT_BATCH_SIZE):
        self.batch_size = batch_size
        pass

    @staticmethod
    def _resolve_output_format(tofile, options: dict) -> str | None:
        """Determine the output data format id from the target path or options.

        Strips a trailing compression-codec extension (e.g. ``.gz``) so that
        ``out.csv.gz`` resolves to ``csv``.
        """
        fmt = options.get("format_out")
        if fmt:
            return fmt.lower()
        basename = str(tofile).rsplit("/", 1)[-1]
        parts = basename.split(".")
        if len(parts) >= 2 and parts[-1].lower() in COMPRESSED_FILE_TYPES:
            parts = parts[:-1]
        if len(parts) >= 2:
            return parts[-1].lower()
        return None

    def _check_output_schema_required(self, tofile, options: dict) -> None:
        """Raise a clear error when the output format needs an external schema.

        Some formats (protobuf, Cap'n Proto, Thrift) report as writable but
        cannot serialize plain records without a compiled schema/message class,
        so they are not valid generic conversion targets.
        """
        out_fmt = self._resolve_output_format(tofile, options)
        if not out_fmt:
            return
        requirement = _SCHEMA_REQUIRED_FORMATS.get(out_fmt)
        if requirement is not None:
            raise ValidationError(
                f"Output format '{out_fmt}' requires {requirement} and cannot be used as a "
                f"generic conversion target. Convert to a self-describing format instead.",
                field="output",
                suggestions=[
                    f for f in WRITABLE_FORMAT_SUGGESTIONS if f not in _SCHEMA_REQUIRED_FORMATS
                ],
            )

    def _check_output_writable(self, tofile, options: dict) -> None:
        """Raise a clear error when the output format is read-only.

        Uses iterabledata's capability reporting so undatum fails fast with an
        actionable message instead of a cryptic engine error.
        """
        out_fmt = self._resolve_output_format(tofile, options)
        if not out_fmt:
            return
        try:
            from iterable.helpers.capabilities import supports_write

            writable = supports_write(out_fmt)
        except Exception:  # noqa: BLE001 - unknown format: let the engine decide
            return
        if writable is False:
            raise ValidationError(
                f"Output format '{out_fmt}' is read-only; undatum can read it but cannot "
                f"write to it.",
                field="output",
                suggestions=WRITABLE_FORMAT_SUGGESTIONS,
            )

    def _build_convert_kwargs(self, options: dict, limit) -> dict:
        """Translate undatum convert options into iterabledata convert kwargs."""
        iterableargs = get_iterable_options(options)
        toiterableargs: dict = {}
        delimiter = options.get("delimiter")
        if delimiter:
            toiterableargs["delimiter"] = delimiter
        is_flatten = bool(get_option(options, "flatten"))
        show_progress = get_option(options, "progress")
        if show_progress is None:
            show_progress = True
        return {
            "iterableargs": iterableargs,
            "toiterableargs": toiterableargs,
            "scan_limit": limit if limit is not None else DEFAULT_HEADERS_DETECT_LIMIT,
            "batch_size": self.batch_size,
            "is_flatten": is_flatten,
            "silent": not show_progress,
            "show_progress": show_progress,
        }

    def convert(self, fromfile, tofile, options=None, limit=DEFAULT_HEADERS_DETECT_LIMIT):
        """Convert a file (or cloud/DB source) to another format.

        Delegates to iterabledata's ``iterable.convert.convert``, which performs
        schema scanning, batched streaming writes, optional flattening, progress
        reporting, atomic local writes, and native cloud (s3/gs/az) handling.
        Friendly errors for missing files and unsupported formats are preserved.

        Args:
            fromfile: Path or URI of the input source.
            tofile: Path or URI of the output file.
            options: Dictionary of conversion options (encoding, delimiter, etc.).
            limit: Maximum records to sample for schema detection.

        Raises:
            FileNotFoundError: If a local input file does not exist.
            PermissionError: If a local file cannot be read.
            FormatError: If the input format is not supported.
        """
        if options is None:
            options = {}

        # Fail fast with an actionable message if the output format is read-only
        # or requires an external schema (protobuf/capnp/thrift).
        self._check_output_writable(tofile, options)
        self._check_output_schema_required(tofile, options)

        # Validate local input files; cloud/DB URIs are validated by the engine.
        if "://" not in fromfile:
            try:
                validate_file_path(fromfile, check_read=True)
            except FileNotFoundError as e:
                suggestions = find_similar_files(fromfile)
                raise FileNotFoundError(fromfile, suggestions) from e
            except PermissionError as e:
                raise PermissionError(fromfile, operation="read") from e

        from iterable.convert import convert as iterable_convert

        convert_kwargs = self._build_convert_kwargs(options, limit)
        # Atomic writes append a ".tmp" suffix to the destination, which breaks
        # format/codec detection for compressed or multi-extension outputs
        # (e.g. "out.csv.gz" -> "out.csv.gz.tmp"). Keep writes non-atomic so the
        # output extension drives detection correctly.

        try:
            iterable_convert(fromfile, tofile, **convert_kwargs)
        except (FileNotFoundError, PermissionError, FormatError):
            raise
        except Exception as e:
            # Surface a helpful error when the input format is not supported.
            file_type = get_file_type(fromfile)
            if file_type is None or file_type not in SUPPORTED_FILE_TYPES:
                raise FormatError(fromfile, file_type, SUPPORTED_FILE_TYPES) from e
            raise

    def bulk_convert(self, source, dest, options=None, to_ext=None, parallel=None):
        """Convert many files (directory or glob) via iterabledata's bulk_convert.

        Args:
            source: Directory path or glob pattern of input files.
            dest: Output directory.
            options: Conversion options (same as :meth:`convert`).
            to_ext: Target extension (e.g. ``"parquet"``); falls back to the
                ``format_out`` option.
            parallel: Force parallel execution; defaults to True when a thread
                count is supplied via options.

        Raises:
            ValueError: If no target extension can be determined.
        """
        if options is None:
            options = {}
        target_ext = to_ext or options.get("format_out")
        if not target_ext:
            raise ValueError(
                "Bulk conversion requires a target extension (use --to-ext or --format-out)."
            )

        # Fail fast if the requested target format is read-only or schema-required.
        self._check_output_writable(f"x.{target_ext}", {})
        self._check_output_schema_required(f"x.{target_ext}", {})

        from iterable.convert import bulk_convert as iterable_bulk_convert

        convert_kwargs = self._build_convert_kwargs(options, DEFAULT_HEADERS_DETECT_LIMIT)
        threads = options.get("threads")
        use_parallel = parallel if parallel is not None else bool(threads)
        return iterable_bulk_convert(
            source,
            dest,
            to_ext=target_ext,
            parallel=use_parallel,
            workers=threads if threads else None,
            **convert_kwargs,
        )
