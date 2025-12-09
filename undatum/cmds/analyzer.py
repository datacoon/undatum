# -*- coding: utf8 -*-
"""Data analysis and insights module.

FIXME: A lot of unoptimized code here, it could be better, shorter and some
functions could be improved
"""
import csv
import io
import json
import os
import tempfile
from collections import OrderedDict
from typing import Optional

import duckdb
import pandas as pd
import xlrd
import xmltodict
import yaml
from iterable.helpers.detect import (detect_encoding_any, detect_file_type,
                                     TEXT_DATA_TYPES)
from openpyxl import load_workbook
from pydantic import BaseModel
from pyzstd import ZstdFile

from ..ai.perplexity import get_fields_info, get_description
from ..formats.docx import analyze_docx
from ..utils import get_dict_value

OBJECTS_ANALYZE_LIMIT = 10000


DUCKDB_TYPES = ['VARCHAR', 'DATE', 'JSON', 'BIGINT', 'DOUBLE', 'BOOLEAN']

def column_type_parse(column_type):
    """Parse column type string to extract array flag and base type."""
    is_array = (column_type[-2:] == '[]')
    if is_array:
        text = column_type[:-2]
    else:
        text = column_type
    if text[:6] == 'STRUCT':
        atype = text[:6]
    elif text[:4] == 'JSON':
        atype = 'VARCHAR'
    else:
        atype = text
    return [atype, str(is_array)]

def duckdb_decompose(filename:str=None, frame:pd.DataFrame=None,
                     filetype:str=None, path:str="*", limit:int=10000000,
                     recursive:bool=True, root:str="", ignore_errors:bool=True):
    """Decomposes file or data frame structure."""
    text_ignore = ', ignore_errors=true' if ignore_errors else ''
    if filetype in ['csv', 'tsv']:
        read_func = f"read_csv('{filename}'{text_ignore})"
    elif filetype in ['json', 'jsonl']:
        read_func = f"read_json('{filename}'{text_ignore})"
    else:
        read_func = f"'{filename}'"
    if path == '*':
        if filename is not None:
            query_str = f"summarize select {path} from {read_func} limit {limit}"
            data = duckdb.sql(query_str).fetchall()
        else:
            query_str = f"summarize select {path} from frame limit {limit}"
            data = duckdb.sql(query_str).fetchall()
    else:
        path_parts = path.split('.')
        query = None
        if len(path_parts) == 1:
            if filename is not None:
                query = (f"summarize select unnest(\"{path}\", recursive:=true) "
                        f"from {read_func} limit {limit}")
            else:
                query = (f"summarize select unnest(\"{path}\", recursive:=true) "
                        f"from frame limit {limit}")
        elif len(path_parts) == 2:
            if filename is not None:
                query = (f"summarize select unnest(\"{path_parts[1]}\", "
                        f"recursive:=true) from (select unnest(\"{path_parts[0]}\", "
                        f"recursive:=true) from {read_func} limit {limit})")
            else:
                query = (f"summarize select unnest(\"{path_parts[1]}\", "
                        f"recursive:=true) from (select unnest(\"{path_parts[0]}\", "
                        f"recursive:=true) from frame limit {limit})")
        elif len(path_parts) == 3:
            if filename is not None:
                query = (f"summarize select unnest(\"{path_parts[2]}\", "
                        f"recursive:=true) from (select unnest(\"{path_parts[1]}\", "
                        f"recursive:=true) from (select unnest(\"{path_parts[0]}\", "
                        f"recursive:=true) from {read_func} limit {limit}))")
            else:
                query = (f"summarize select unnest(\"{path_parts[2]}\", "
                        f"recursive:=true) from (select unnest(\"{path_parts[1]}\", "
                        f"recursive:=true) from (select unnest(\"{path_parts[0]}\", "
                        f"recursive:=true) from frame limit {limit}))")
        elif len(path_parts) == 4:
            if filename is not None:
                query = (f"summarize select unnest(\"{path_parts[2]}.{path_parts[3]}\", "
                        f"recursive:=true) from (select unnest(\"{path_parts[1]}\", "
                        f"recursive:=true) from (select unnest(\"{path_parts[0]}\", "
                        f"recursive:=true) from {read_func} limit {limit}))")
            else:
                query = (f"summarize select unnest(\"{path_parts[2]}.{path_parts[3]}\", "
                        f"recursive:=true) from (select unnest(\"{path_parts[1]}\", "
                        f"recursive:=true) from (select unnest(\"{path_parts[0]}\", "
                        f"recursive:=true) from frame limit {limit}))")
        data = duckdb.sql(query).fetchall()
    table = []
    for row in data:
        item = [row[0] if len(root) == 0 else root + '.' + row[0]]
        item.extend(column_type_parse(row[1]))
        item.append(str(row[4]))
        item.append(str(row[10]))
        uniq_share = row[4] * 100.0 / row[10] if row[10] > 0 else 0
        item.append(f'{uniq_share:0.2f}')
        table.append(item)
        if recursive and item[1] == 'STRUCT':
            sub_path = row[0] if len(root) == 0 else item[0]
            subtable = duckdb_decompose(filename, frame, filetype=filetype,
                                       path=sub_path, limit=limit,
                                       recursive=recursive, root=item[0],
                                       ignore_errors=ignore_errors)
            for subitem in subtable:
                table.append(subitem)
    return table

def _seek_dict_lists(data, level=0, path=None, candidates=None):
    """Seek list structures in dictionary recursively."""
    if candidates is None:
        candidates = OrderedDict()
#    print(level, path, candidates)
    for key, value in data.items():
        if isinstance(value, list):
            isobjectlist = False
            for listitem in value[:20]:
                if isinstance(listitem, (dict, OrderedDict)):
                    isobjectlist = True
                    break
            if not isobjectlist:
                continue
            key = f'{path}.{key}' if path is not None else key
            if key not in candidates:
                candidates[key] = {'key' : key, 'num' : len(value)}
        elif isinstance(value, (OrderedDict, dict)):
            res = _seek_dict_lists(value, level + 1, path + '.' + key if path else key, candidates)
            for k, v in res.items():
                if k not in candidates.keys():
                    candidates[k] = v
        else:
            continue
    return candidates



def _seek_xml_lists(data, level=0, path=None, candidates=None):
    """Seek list structures in XML data recursively."""
    if candidates is None:
        candidates = OrderedDict()
    for key, value in data.items():
        if isinstance(value, list):
            key = f'{path}.{key}' if path is not None else key
            if key not in candidates:
                candidates[key] = {'key' : key, 'num' : len(value)}
        elif isinstance(value, (OrderedDict, dict)):
            res = _seek_xml_lists(value, level + 1, path + '.' + key if path else key, candidates)
            for k, v in res.items():
                if k not in candidates.keys():
                    candidates[k] = v
        else:
            continue
    return candidates


def _process_json_data(data, report, fullkey, objects_limit, use_pandas,
                      autodoc, lang):
    """Process JSON data and add tables to report."""
    candidates = _seek_dict_lists(data, level=0)
    if len(candidates) == 1:
        fullkey = str(next(iter(candidates)))
        table = TableSchema(id=fullkey)
        objects = get_dict_value(data, keys=fullkey.split('.'))[0]
        table = table_from_objects(objects, table_id=fullkey,
                                  objects_limit=objects_limit,
                                  use_pandas=use_pandas,
                                  filetype='jsonl',
                                  autodoc=autodoc, lang=lang)
        report.tables.append(table)
        report.total_tables = len(report.tables)
        report.total_records = table.num_records
    elif len(candidates) > 1:
        total = 0
        for fullkey in candidates:
            table = TableSchema(id=fullkey)
            objects = get_dict_value(data, keys=fullkey.split('.'))[0]
            table = table_from_objects(objects, table_id=fullkey,
                                      objects_limit=objects_limit,
                                      use_pandas=use_pandas,
                                      filetype='jsonl',
                                      autodoc=autodoc, lang=lang)
            total += table.num_records
            report.tables.append(table)
        report.total_records = total
        report.total_tables = len(report.tables)


class FieldSchema(BaseModel):
    """Schema definition for a data field."""
    name: str
    ftype: str
    is_array:bool = False
    description: Optional[str] = None
    sem_type:str = None
    sem_url:str = None


class TableSchema(BaseModel):
    """Table schema definition."""
    num_records: int = -1
    num_cols: int = -1
    is_flat: bool = True
    id: Optional[str] = None
    fields: Optional[list[FieldSchema]] = []
    description: Optional[str] = None

class ReportSchema(BaseModel):
    """Schema of the data file analysis results."""
    filename: str
    file_size: int
    file_type: str
    compression: str = None
    total_tables: int = 1
    total_records: int = -1
    tables: Optional[list[TableSchema]] = []
    metadata: dict = {}
    success: bool = False
    error: str = None


MAX_SAMPLE_SIZE = 200
DELIMITED_FILES = ['csv', 'tsv']
DUCKABLE_FILE_TYPES = ['csv', 'jsonl', 'json', 'parquet']
DUCKABLE_CODECS = ['zst', 'gzip', 'raw']


def table_from_objects(objects: list, table_id: str, objects_limit: int,
                      use_pandas: bool = False, filetype='csv',
                      autodoc: bool = False, lang: str = 'English'):
    """Reconstructs table schema from list of objects."""
    table = TableSchema(id=table_id)
    table.num_records = len(objects)
    if autodoc:
        f = io.StringIO()
        writer = csv.writer(f)
        writer.writerows(objects[:MAX_SAMPLE_SIZE])
        table.description = get_description(f.getvalue(), language=lang)
    if use_pandas:
        df = pd.DataFrame(objects)
        columns_raw = duckdb_decompose(frame=df, path='*',
                                      limit=objects_limit)
    else:
        suffix = '.' + filetype
        tfile = tempfile.NamedTemporaryFile(suffix=suffix, mode='w',
                                            encoding='utf8', delete=False)
        tfile.close()
        with ZstdFile(tfile.name, mode='w', level_or_option=9) as tfile_real:
            wrapper = io.TextIOWrapper(tfile_real, encoding='utf8',
                                       write_through=True)
            if filetype == 'csv':
                writer = csv.writer(wrapper)
                writer.writerows(objects[:objects_limit])
            elif filetype == 'jsonl':
                for row in objects[:objects_limit]:
                    wrapper.write(json.dumps(row) + '\n')
        # Getting structure
        columns_raw = duckdb_decompose(tfile.name, filetype=filetype,
                                       path='*', limit=objects_limit)
        os.remove(tfile.name)
    is_flat = True
    table.num_cols = len(columns_raw)

    for column in columns_raw:
        field = FieldSchema(name=column[0], ftype=column[1],
                           is_array=column[2])
        table.fields.append(field)
        if field.ftype == 'STRUCT' or field.is_array:
            is_flat = False
        table.is_flat = is_flat
    table.num_records = len(objects)
    return table




def analyze(filename: str, filetype: str = None, compression: str = 'raw',
           objects_limit: int = OBJECTS_ANALYZE_LIMIT, encoding: str = None,
           scan: bool = True, stats: bool = True, engine: str = "auto",  # noqa: ARG001
           use_pandas: bool = False, ignore_errors: bool = True,
           autodoc: bool = False, lang: str = 'English'):
    """Analyzes any type of data file and provides meaningful insights."""
    fileext = filename.rsplit('.', 1)[-1].lower()
    filesize = os.path.getsize(filename)
    if filetype is None:
        ftype = detect_file_type(filename)
        if ftype['success']:
            filetype = ftype['datatype'].id()
            if ftype['codec'] is not None:
                compression = ftype['codec'].id()
    # Handling special cases
    if filetype is None and fileext == 'docx':
        filetype = 'docx'

    report = ReportSchema(filename=filename, file_size=filesize,
                         file_type=filetype, compression=compression)
    if filetype in TEXT_DATA_TYPES:
        if encoding is None:
            encoding = detect_encoding_any(filename)
            enc_key = 'encoding' if 'encoding' in encoding else None
            report.metadata['encoding'] = encoding.get(enc_key) if enc_key else None
        else:
            report.metadata['encoding'] = encoding
    if scan:
        duckable_cond = (report.file_type in DUCKABLE_FILE_TYPES and
                        report.compression in DUCKABLE_CODECS and
                        engine in ['auto', 'duckdb'])
        if duckable_cond:
            # Getting total count
            text_ignore = ', ignore_errors=true' if ignore_errors else ''
            if filetype in ['json', 'jsonl']:
                query_str = f"select count(*) from read_json('{filename}'{text_ignore})"
                num_records = duckdb.sql(query_str).fetchall()[0][0]
            elif filetype in ['csv', 'tsv']:
                query_str = f"select count(*) from read_csv('{filename}'{text_ignore})"
                num_records = duckdb.sql(query_str).fetchall()[0][0]
            else:
                query_str = f"select count(*) from '{filename}'"
                num_records = duckdb.sql(query_str).fetchall()[0][0]
            table = TableSchema(id=os.path.basename(filename))
            table.num_records = num_records
            report.tables = [table]
            report.total_records = table.num_records
            report.total_tables = 1

            # Getting structure
            columns_raw = duckdb_decompose(filename, filetype=filetype,
                                          path='*', limit=objects_limit)
            is_flat = True
            table.num_cols = len(columns_raw)
            for column in columns_raw:
                field = FieldSchema(name=column[0], ftype=column[1],
                                   is_array=column[2])
                table.fields.append(field)
                if field.ftype == 'STRUCT' or field.is_array:
                    is_flat = False
            table.is_flat = is_flat
            query_str = f"select * from '{filename}' limit {MAX_SAMPLE_SIZE}"
            sample = duckdb.sql(query_str).fetchall()
            f = io.StringIO()
            writer = csv.writer(f)
            writer.writerows(sample[:MAX_SAMPLE_SIZE])
            if autodoc:
                table.description = get_description(f.getvalue(), language=lang)
        else:
            if engine == 'duckdb':
                report.success = False
                report.error = (f"Not supported file type {report.file_type} "
                               f"or compression {report.compression}")
            else:
                # Processing MS Word XML files
                if fileext == 'docx':
                    docx_tables = analyze_docx(filename, extract_data=True)
                    total = 0
                    for dtable in docx_tables:
                        table = table_from_objects(dtable['data'],
                                                   table_id=str(dtable['id']),
                                                   objects_limit=objects_limit,
                                                   use_pandas=use_pandas,
                                                   filetype='csv',
                                                   autodoc=autodoc, lang=lang)
                        total += table.num_records
                        report.tables.append(table)
                    report.total_records = total
                    report.total_tables = len(report.tables)
                elif filetype == 'xlsx':
                    wb = load_workbook(filename)
                    total = 0
                    for sheetname in wb.sheetnames:
                        sheet = wb.get_sheet_by_name(sheetname)
                        objects = []
                        max_num = (objects_limit if objects_limit < sheet.max_row
                                  else sheet.max_row)
                        for n in range(0, max_num):
                            row = next(sheet.iter_rows())
                            tmp = []
                            for cell in row:
                                tmp.append(str(cell.value))
                            objects.append(tmp)
                        table = table_from_objects(objects, table_id=sheetname,
                                                   objects_limit=objects_limit,
                                                   use_pandas=use_pandas,
                                                   filetype='csv',
                                                   autodoc=autodoc, lang=lang)
                        total += table.num_records
                        report.tables.append(table)
                    report.total_records = total
                    report.total_tables = len(report.tables)
                elif filetype == 'xls':
                    wb = xlrd.open_workbook(filename)
                    total = 0
                    for sheetname in wb.sheet_names():
                        sheet = wb.sheet_by_name(sheetname)
                        objects = []
                        max_num = (objects_limit if objects_limit < sheet.nrows
                                  else sheet.nrows)
                        for n in range(0, max_num):
                            tmp = []
                            for i in range(0, sheet.ncols):
                                cell_value = sheet.cell_value(n, i)
                                get_col = str(cell_value)
                                tmp.append(get_col)
                            objects.append(tmp)
                        table = table_from_objects(objects, table_id=sheetname,
                                                  objects_limit=objects_limit,
                                                  use_pandas=use_pandas,
                                                  filetype='csv',
                                                  autodoc=autodoc, lang=lang)
                        report.tables.append(table)
                        total += table.num_records
                    report.total_records = total
                    report.total_tables = len(report.tables)
                elif filetype == 'xml':
                    fileobj = None
                    codec = None
                    if ftype['codec'] is not None:
                        codec = ftype['codec'](filename, open_it=True)
                        fileobj = codec.fileobj()
                    if fileobj is None:
                        with open(filename, 'rb') as f:
                            data = xmltodict.parse(f, process_namespaces=False)
                    else:
                        data = xmltodict.parse(fileobj, process_namespaces=False)
                    candidates = _seek_xml_lists(data, level=0)
                    if len(candidates) == 1:
                        fullkey = str(next(iter(candidates)))
                        table = TableSchema(id=fullkey)
                        objects = get_dict_value(data,
                                                keys=fullkey.split('.'))[0]
                        table = table_from_objects(objects, table_id=fullkey,
                                                  objects_limit=objects_limit,
                                                  use_pandas=use_pandas,
                                                  filetype='jsonl',
                                                  autodoc=autodoc, lang=lang)
                        report.tables.append(table)
                        report.total_tables = len(report.tables)
                        report.total_records = table.num_records
                    elif len(candidates) > 1:
                        total = 0
                        for fullkey in candidates:
                            table = TableSchema(id=fullkey)
                            objects = get_dict_value(data,
                                                    keys=fullkey.split('.'))[0]
                            table = table_from_objects(objects, table_id=fullkey,
                                                      objects_limit=objects_limit,
                                                      use_pandas=use_pandas,
                                                      filetype='jsonl',
                                                      autodoc=autodoc, lang=lang)
                            total += table.num_records
                            report.tables.append(table)
                        report.total_records = total
                        report.total_tables = len(report.tables)
                    if codec is not None:
                        codec.close()
                    else:
                        fileobj.close()
                elif filetype == 'json':
                    fileobj = None
                    codec = None
                    if ftype['codec'] is not None:
                        codec = ftype['codec'](filename, open_it=True)
                        fileobj = codec.fileobj()
                    if fileobj is None:
                        with open(filename, 'rb') as f:
                            data = json.load(f)
                    else:
                        data = json.load(fileobj)
                    _process_json_data(data, report, fullkey,
                                     objects_limit, use_pandas,
                                     autodoc, lang)
                    if codec is not None:
                        codec.close()
                    elif fileobj is not None:
                        fileobj.close()

    if autodoc and report.total_tables > 0:
        for table in report.tables:
            fields = []
            for column in table.fields:
                fields.append(column.name)
            descriptions = get_fields_info(fields, language=lang)
            for column in table.fields:
                if column.name in descriptions:
                    column.description = descriptions[column.name]
    return report





class Analyzer:
    """Data analysis handler."""
    def __init__(self):
        pass


    def analyze(self, filename, options):
        """Analyzes given data file and returns it's parameters"""
        from tabulate import tabulate

        table = None
        encoding = options.get('encoding')
        report = analyze(filename, encoding=encoding,
                        engine=options['engine'],
                        use_pandas=options['use_pandas'],
                        autodoc=options['autodoc'], lang=options['lang']) 
        if options['outtype'] == 'json':
            if options['output'] is not None:
                with open(options['output'], 'w', encoding='utf8') as f:
                    f.write(json.dumps(report.model_dump()))
            else:
                print(json.dumps(report.model_dump(), indent=4, ensure_ascii=False))
        if options['outtype'] == 'yaml': 
            if options['output'] is not None:
                with open(options['output'], 'w', encoding='utf8') as f:
                    f.write(yaml.dump(report.model_dump(), Dumper=yaml.Dumper))
            else:
                print(yaml.dump(report.model_dump(), Dumper=yaml.Dumper))

        elif options['outtype'] == 'markdown':
            raise NotImplementedError("Markdown output not implemented")
        else:
            print('Analysis report')
            headers = ['Attribute', 'Value']
            reptable = []
            reptable.append(['Filename', str(report.filename)])
            reptable.append(['File size', str(report.file_size)])
            reptable.append(['File type', report.file_type])
            reptable.append(['Compression', str(report.compression)])
            reptable.append(['Total tables', str(report.total_tables)])
            reptable.append(['Total records', str(report.total_records)])
            for k, v in report.metadata.items():
                reptable.append([k, v])
            print(tabulate(reptable, headers=headers))

            tabheaders = ['Name', 'Type', 'Is Array', 'Description']
            for rtable in report.tables:
                print()
                table = []
                msg = (f"Table {rtable.id} (items {rtable.num_records}, "
                      f"columns {rtable.num_cols}) structure")
                print(msg)
                for field in rtable.fields:
                    table.append([field.name, field.ftype, str(field.is_array),
                                 field.description])
                print(tabulate(table, headers=tabheaders))
                print("Summary:")
                print(rtable.description)                
                